
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SAMPLE_DIR=Path("sample_data")
SAMPLE_DIR.mkdir(exist_ok=True)

def style(ws):
    fill=PatternFill("solid",fgColor="1F2937")
    font=Font(color="FFFFFF",bold=True)
    side=Side(style="thin",color="CBD5E1")
    border=Border(left=side,right=side,top=side,bottom=side)
    for c in ws[1]:
        c.fill=fill; c.font=font; c.alignment=Alignment(horizontal="center"); c.border=border
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border=border; c.alignment=Alignment(horizontal="center")
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=max(len(str(c.value)) if c.value else 0 for c in col)+4
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions

def save(name, title, rows):
    wb=Workbook(); ws=wb.active; ws.title=title
    for r in rows: ws.append(r)
    style(ws); wb.save(SAMPLE_DIR/name)

save("DataTalk_工程測試資料.xlsx","工程資料",[
["工單編號","日期","部門","區域","狀態","供應商","負責人","數量","備註"],
["WO001","2026-06-01","工程部","A區","完成","華新工程","小明",12,"正常"],
["WO002","2026-06-01","工程部","B區","異常","華新工程","小華",3,"材料缺失"],
["WO003","2026-06-02","品保部","A區","完成","宏達科技","小美",8,"正常"],
["WO004","2026-06-02","製造部","C區","延遲","建安企業","阿豪",5,"設備延遲"],
["WO005","2026-06-03","工程部","A區","NG","華新工程","小明",2,"施工缺失"],
["WO006","2026-06-03","品保部","B區","完成","宏達科技","小美",9,"正常"],
["WO007","2026-06-04","製造部","C區","待處理","建安企業","阿豪",4,"待主管確認"],
["WO008","2026-06-04","工程部","A區","完成","華新工程","小華",11,"正常"],
["WO009","2026-06-05","品保部","B區","異常","宏達科技","小美",1,"檢查未通過"],
["WO010","2026-06-05","製造部","C區","完成","建安企業","阿豪",7,"正常"],
])
save("DataTalk_供應商合併資料.xlsx","供應商資料",[
["供應商","聯絡人","電話","等級","付款條件"],
["華新工程","王先生","0911-111-111","A","月結30天"],
["宏達科技","林小姐","0922-222-222","B","月結45天"],
["建安企業","陳先生","0933-333-333","A","月結30天"],
])
save("DataTalk_修改功能測試.xlsx","修改測試",[
["工單編號","狀態","備註"],
["WO001","完成","不用修改"],
["WO002","未完成","測試取代文字"],
["WO003",None,"測試補空白"],
["WO004","異常","測試篩選"],
])
print("sample_data 已產生完成")
