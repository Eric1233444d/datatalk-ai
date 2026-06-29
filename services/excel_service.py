import io
from pathlib import Path
from copy import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SAMPLE_DIR = Path(__file__).resolve().parents[1] / "sample_data"


def _score_df(df: pd.DataFrame) -> float:
    if df is None or df.empty:
        return -1
    cols = [str(c).strip() for c in df.columns]
    non_empty_cols = sum(1 for c in cols if c and not c.lower().startswith("unnamed"))
    non_empty_cells = int(df.notna().sum().sum())
    unique_cols = len(set(cols))
    duplicate_penalty = len(cols) - unique_cols
    unnamed_penalty = sum(1 for c in cols if c.lower().startswith("unnamed"))
    return non_empty_cols * 20 + non_empty_cells - duplicate_penalty * 8 - unnamed_penalty * 6


def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    # remove fully empty rows/cols
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    # normalize column names and make duplicates unique
    new_cols = []
    seen = {}
    for i, c in enumerate(df.columns):
        name = str(c).strip()
        if not name or name.lower().startswith("unnamed"):
            name = f"欄位{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        new_cols.append(name)
    df.columns = new_cols
    return df.reset_index(drop=True)


def _read_best_sheet(data: bytes):
    xls = pd.ExcelFile(io.BytesIO(data))
    candidates = []
    for sheet in xls.sheet_names:
        for header in range(0, 12):
            try:
                df = pd.read_excel(io.BytesIO(data), sheet_name=sheet, header=header)
                df = _clean_df(df)
                candidates.append(( _score_df(df), sheet, header, df))
            except Exception:
                pass
    if not candidates:
        df = pd.read_excel(io.BytesIO(data))
        return _clean_df(df), {"sheet": None, "header_row": 1, "all_sheets": []}
    candidates.sort(key=lambda x: x[0], reverse=True)
    _, sheet, header, df = candidates[0]
    return df, {"sheet": sheet, "header_row": header + 1, "all_sheets": xls.sheet_names}


def read_excel(uploaded_file):
    data = uploaded_file.getvalue()
    df, meta = _read_best_sheet(data)
    return df, data, meta


def read_excel_bytes(data: bytes):
    return _read_best_sheet(data)


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="DataTalk_Result")
        ws = writer.book["DataTalk_Result"]
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            letter = col[0].column_letter
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
            ws.column_dimensions[letter].width = min(max_len + 4, 45)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    return output.getvalue()


def preserve_style_export(original_bytes: bytes | None, df: pd.DataFrame) -> bytes:
    if not original_bytes:
        return dataframe_to_excel_bytes(df)
    try:
        wb = load_workbook(io.BytesIO(original_bytes))
        ws = wb.active
        # Clear existing values conservatively, then rewrite result
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None
        for c_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=c_idx).value = col_name
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = None if pd.isna(value) else value
                if r_idx > 2:
                    src = ws.cell(row=2, column=c_idx)
                    if src.has_style:
                        cell._style = copy(src._style)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    except Exception:
        return dataframe_to_excel_bytes(df)


def get_sample_files():
    return sorted(SAMPLE_DIR.glob("*.xlsx"))
